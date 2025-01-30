import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import random

OS_Windows = False
if os.name == 'nt':  # the code is running on windows
    OS_Windows = True
if OS_Windows:
    TESTS_MODELS = '/media/rasho/M2_Samsung990/Work/Models/EarlyBird/tests'
else:
    TESTS_MODELS = '/media/rasho/M2_Samsung990/Work/Models/EarlyBird/tests'

def get_result_files_path():
    list_of_pathes = []
    list_of_tests = []
    for folder in os.listdir(TESTS_MODELS):
        if folder.startswith('Assessing_'):
            test_name = '_'.join(folder.split('_', 1)[1:])
            list_of_tests.append(test_name)
            list_of_pathes.append(os.path.join(TESTS_MODELS, folder, 'console_output.txt'))

    return list_of_pathes, list_of_tests


def split_tests(test_list):
    tests_splits = []
    for test in test_list:
        parts = test.split('_')
        tests_splits.append('_'.join(parts[:2]))
    test_split_set = set(tests_splits)

    color_list = generate_unique_colors(len(test_split_set))
    print(color_list)
    split_test_color = {key: value for key, value in zip(test_split_set, color_list)}
    return split_test_color


def generate_unique_colors(n):
    COLOR_INDEX = (
        # '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
        '00FFFF00', '00FF00FF', '0000FFFF', '00FFFFFF',  # 5-9
        '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
        '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
        '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
        '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
        '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
        '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
        '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
        '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
        '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
        '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
        '00993300', '00993366', '00333399', '00333333',  # 60-63
    )

    colors = COLOR_INDEX[:n]
    return colors


# List to hold data from all files
all_data = []

simplified_list = [  # 'detect/mAP_3D',
    'detect/moda', 'detect/modp',
    'detect/precision', 'detect/recall', 'detect/mAP',
    'track/idf1',
    # 'track/idp', 'track/idr',
    'track/mota', 'track/motp'
]
simplified_data = []

# Iterate through each file in the directory
tests_path, tests_names = get_result_files_path()

for filename in tests_path:  # [directory]:  # os.listdir(directory):
    if filename.endswith('.txt'):
        # filepath = os.path.join(directory, filename)
        filepath = filename
        with open(filepath, 'r') as file:
            lines = file.readlines()

        # Extract the test name from the file
        test_name = lines[1].strip().replace('-----', '').strip()
        test_name = '_'.join(test_name.split('_', 1)[1:])

        # Extract metrics
        metrics = {}
        simplified_metrics = {}
        for line in lines:
            if '│' in line and 'detect' in line or 'track' in line:
                parts = line.split('│')
                if len(parts) > 2:
                    metric_name = parts[1].strip()
                    metric_value = parts[2].strip()
                    metrics[metric_name] = round(float(metric_value), 3)
                    # print(metric_name)
                    if metric_name in simplified_list:
                        simplified_metrics[metric_name] = round(float(metric_value), 3)

        # Add the test name to the metrics dictionary
        metrics['Test Name'] = test_name
        simplified_metrics['Test Name'] = test_name

        all_data.append(metrics)
        simplified_data.append(simplified_metrics)

# Create a DataFrame from the list of dictionaries
df = pd.DataFrame(all_data)
simplified_df = pd.DataFrame(simplified_data)

# Rearrange columns to have 'Test Name' as the first column
columns = ['Test Name'] + [col for col in df.columns if col != 'Test Name']
df = df[columns]

columns = ['Test Name'] + [col for col in simplified_df.columns if col != 'Test Name']
simplified_df = simplified_df[columns]

df = df.sort_values(by='Test Name')
simplified_df = simplified_df.sort_values(by='Test Name')

# Save the DataFrame to a CSV file
simplified_df.to_csv('combined_metrics.csv', index=False)

excel_file = 'util_outputs/simplified_df.xlsx'
simplified_df.to_excel(excel_file, engine='openpyxl')

wb = load_workbook(excel_file)
ws = wb.active  # Select the active worksheet

test_name_color = split_tests(tests_names)

# Iterate through rows and highlight based on 'Test Name' column
for row in ws.iter_rows(min_row=2, max_row=len(simplified_df) + 1, min_col=1, max_col=len(simplified_df.columns) + 1):
    cell_value = row[1].value  # First column (Test Name)
    test_name_keys = test_name_color.keys()
    for test_name_key in test_name_keys:
        if cell_value.startswith(test_name_key):
            use_color = test_name_color[test_name_key]
            fill = PatternFill(start_color=use_color, end_color=use_color, fill_type='solid')  # Yellow color
            for cell in row:
                cell.fill = fill

######
###### beutify the table

# Write headers to Excel
for col_num, column_title in enumerate(simplified_df.columns, start=1):
    ws.cell(row=1, column=col_num, value=column_title)

# Write data to Excel
for row_num, row_data in enumerate(simplified_df.values, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Adjust column widths to fit content
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

# Add borders to cells
for row in ws.iter_rows(min_row=1, max_row=len(simplified_df) + 1, min_col=1, max_col=len(simplified_df.columns) + 1):
    for cell in row:
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        cell.border = border
########
########


# Save workbook with changes
wb.save(excel_file)
